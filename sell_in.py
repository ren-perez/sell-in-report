# %%

import os
# import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# %%

# Load data
response = pd.read_excel("sell_in_mod.xlsx", sheet_name="LIMA", skiprows=4)
zonificacion = pd.read_excel("sell_in_mod.xlsx", sheet_name="Listas-clientes")
ubicaciones = pd.read_excel("sell_in_mod.xlsx", sheet_name="Listas-ubicaciones")
articulos = pd.read_excel("sell_in_mod.xlsx", sheet_name="Listas-articulos")

ubicaciones = ubicaciones.drop_duplicates().reset_index(drop=True)
articulos = articulos.drop_duplicates().reset_index(drop=True)

# %%

# Filter only NC_Ventas, Ventas
df = response[response["TipoVenta"].isin(["NC_Ventas", "Ventas"])]
df = df[df["Emision"].dt.year >= 2023]

# Convert specified columns to uppercase
df['ProvinciaDespacho'] = df['ProvinciaDespacho'].str.upper()
df['DistritoDespacho'] = df['DistritoDespacho'].str.upper()

# Filter the relevant columns
cols = ['Emision', 'ProvinciaDespacho', 'CodCliente', 'Cliente', 'CodigoPago', 'ListaPrecio', 'CodArticulo', 'Articulo', 'Caja 9L', 'ValorVenta', 'CantidadDespachada', 'CostoVentas']
# cols = ['Emision', 'ProvinciaDespacho', 'CodCliente', 'Cliente', 'Caja 9L', 'ValorVenta', 'CantidadDespachada', 'CostoVentas']
df = df[cols]

# %%

# Group by the specified columns and sum the relevant columns
grouped_df = df.groupby(['Emision', 'ProvinciaDespacho', 'CodCliente', 'Cliente', 'CodigoPago', 'ListaPrecio', 'CodArticulo', 'Articulo'], dropna=False).sum().reset_index()

# Group by 'CodCliente' and fill NaN in 'ProvinciaDespacho' with the group's value
grouped_df['ProvinciaDespacho'] = grouped_df.groupby('CodCliente')['ProvinciaDespacho'].transform(lambda x: x.fillna(method='ffill').fillna(method='bfill'))

# grouped_df

# %%

# SHEET SELL IN

grouped_df['Suma de Caja 9L'] = grouped_df['Caja 9L']
grouped_df['Suma de ValorVenta'] = grouped_df['ValorVenta']
grouped_df['Suma de CantidadDespachada'] = grouped_df['CantidadDespachada']
grouped_df['Suma de CantidadDespachada2'] = grouped_df['CostoVentas']

# Join Ubicaciones
sell_in = pd.merge(grouped_df, ubicaciones[['Ubicación BD', 'Ubicación']], how='left', left_on='ProvinciaDespacho', right_on='Ubicación BD')

# Fill missing values in the join column with 'Otros'
sell_in['Ubicación'] = sell_in['Ubicación'].fillna('Otros', inplace=False)

# Column Ubicacion
sell_in['Concatenado (cliente SAP-ubi)'] = sell_in['Cliente'] + sell_in['Ubicación']

# Join Zonificaion
sell_in = pd.merge(sell_in, zonificacion[['Concatenado', 'Canal']], how='left', left_on='Concatenado (cliente SAP-ubi)', right_on='Concatenado')

# Fill missing values in the join column with 'Otros'
sell_in['Canal'] = sell_in['Canal'].fillna('Otros', inplace=False)

# Join Articulos (Tabla4)
sell_in = pd.merge(sell_in, articulos, how='left', left_on='Articulo', right_on='Articulo')

# Fill missing values in the join column with 'Otros'
sell_in['Articulo (Sin duplicados)'] = sell_in['Articulo 2']

sell_in = sell_in[["Emision", "ProvinciaDespacho", "Canal", "Ubicación", "CodCliente", "Cliente", 'Concatenado (cliente SAP-ubi)', "CodigoPago", "ListaPrecio", "CodArticulo", "Articulo", "Articulo (Sin duplicados)", "Suma de Caja 9L", "Suma de ValorVenta", "Suma de CantidadDespachada", "Suma de CantidadDespachada2"]]

# sell_in

#%%

# Save the result to a new Excel file
sell_in.to_excel("Reporte de Sell in (Actualizable).xlsx", index=False)


# # Create a new workbook and select the active sheet
# wb = Workbook()
# ws = wb.active

# # Write the DataFrame to the worksheet
# for r_idx, row in enumerate(dataframe_to_rows(sell_in, index=False, header=True), 1):
#     for c_idx, value in enumerate(row, 1):
#         ws.cell(row=r_idx, column=c_idx, value=value)

# # Define styles
# header_font = Font(bold=True, color="FFFFFF")
# blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
# brown_fill = PatternFill(start_color="833C0C", end_color="833C0C", fill_type="solid")
# alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
# thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# # Apply formatting to header row
# for cell in ws[1]:
#     cell.font = header_font
#     cell.border = thin_border
#     cell.alignment = Alignment(horizontal='center', vertical='center')
    
#     # Apply brown fill to specific columns
#     if cell.column_letter in ['C', 'D', 'G']:
#         cell.fill = brown_fill
#     else:
#         cell.fill = blue_fill

# # Apply alternating row colors and borders to data rows
# for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
#     for cell in row:
#         cell.border = thin_border
#         if row[0].row % 2 == 0:
#             cell.fill = alt_row_fill

# # Adjust column widths
# for column in ws.columns:
#     max_length = 0
#     column_letter = column[0].column_letter
#     for cell in column:
#         try:
#             if len(str(cell.value)) > max_length:
#                 max_length = len(cell.value)
#         except:
#             pass
#     adjusted_width = (max_length + 2) * 1.2
#     ws.column_dimensions[column_letter].width = adjusted_width

# # Save the workbook
# filename = "Reporte de Sell in (Actualizable).xlsx"
# try:
#     if os.path.exists(filename):
#         os.remove(filename)  # Delete the file if it already exists
#     wb.save(filename)
#     print(f"File {filename} has been successfully created/overwritten.")
# except PermissionError:
#     print(f"Unable to save the file. Please close {filename} if it's open in Excel.")
# except Exception as e:
#     print(f"An error occurred while saving the file: {e}")

# %%

# Define the connection string
conn_str = (
    'DRIVER={SQL Server};'
    'SERVER=192.168.0.209;'
    'UID=sa;'
    'PWD=SAPB1Admin;'
    'APP=Microsoft Office 2016;'
    'WSID=BP-PC80;'
    'DATABASE=SBO_BEBIDASPREMIUM'
)

# Connect to the database
conn = pyodbc.connect(conn_str)

# Define the SQL query
sql_query = """
SELECT TipoVenta, DocNum, NumeroDocumento, Emision, CodCliente, Cliente, CodigoPago, Segmento,
       CodArticulo, Articulo, Paquete, Vendedor, Moneda, TC, Precio, PorDscto, Dscto, ValorVenta, 
       ISCtasa, IGV, TotalIclIGV, DepartamentoDespacho, ProvinciaDespacho, DistritoDespacho, 
       DireccionDespacho, DeparamentoFiscal, ProvinciaFiscal, DistritoFiscal, DireccionFiscal, 
       ListaPrecio, CostoVentas, NombreComercial, GrupoComercial, asiento, IndicadorImpuestos, 
       CuentaMayor, DocumentoBase, CuentaCosto, Item, Fecha_de_vencimiento, Tipo_de_Seguro, 
       CantidadDespachada, Tipo_Bonif_1, Cantidad_Bonif_1, Articulo_Bonif_2, Tipo_Bonif_2, 
       Cantidad_Bonif_2, Factor_Conversión_LT, ArtxUnidadxCaja, Factor, CajaFísica, Caja_9L, 
       Canal, Zona, Marca, CostoVentas1, CostoVentas2, destino, Comentario, Tipo_Documento, 
       Tipo_Venta, Orden_de_Fabricación, Cantidad_Producida, Cod_Art_Consumida, 
       Nom_Art_Consumida, Cant_Consumida, Costo_Unit_Consumido, Costo_Total_Consumido, 
       Marca_Consumida, RefVenta_Consumido, ISC_Consumido, IGV_Consumido, PV_Full_Consumido, 
       RUC, Nombre_Fam, Nombre_Sub_Fam, Ejecutvo_asignado, NºSAPDocumentoBase, 
       DocumentoBase_FechadeEntrega, DocumentoBase_ModalidadTraslado, DocumentoBase_CodigoTransportista, 
       DocumentoBase_RUCTransportista, DocumentoBase_NombreTransportista
FROM YourTableName  -- Replace with your actual table name
WHERE YourCondition  -- Replace with your actual condition if any
"""

# Execute the SQL query
data = pd.read_sql(sql_query, conn)

# # Define the stored procedure and parameters
# stored_procedure = '{Call BPL_SP_Detalle_Ventas(?,?)}'
# params = (param1, param2)  # Replace with actual parameters

# # Execute the stored procedure
# cursor = conn.cursor()
# cursor.execute(stored_procedure, params)

# # Fetch the results
# columns = [column[0] for column in cursor.description]
# response = cursor.fetchall()

# # Convert the results to a DataFrame
# response = pd.DataFrame.from_records(response, columns=columns)

# Close the connection
conn.close()
















# %%
# cases
check = grouped_df[(grouped_df['Emision'].dt.year == 2024) & (grouped_df['Emision'].dt.month == 2)]
check[(check['ProvinciaDespacho'] == 'LIMA') & (check['Emision'].dt.day == 12) & (check['CodCliente'] == 'CL20603269439')]